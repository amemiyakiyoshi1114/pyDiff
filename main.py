# This is a sample Python script.
import sys

import openpyxl
import pandas
from PyQt5 import QtWidgets
from PIL import Image
from openpyxl.reader.excel import load_workbook

from ui.diffPage import UiDiffPage


# Press Shift+F10 to execute it or replace it with your code.
# Press Double Shift to search everywhere for classes, files, tool windows, actions, and settings.

def find_type():
    df = pandas.read_excel("D:\\Users\\zhongsiyang\\test.xlsx")
    wb = load_workbook("D:\\Users\\zhongsiyang\\test.xlsx")
    ws = wb[wb.sheetnames[0]]
    img1=[]
    img2=[]
    img3=[]
    i = 0
    for image in ws._images:
        i += 1
        # 将图片转换成Pillow中的图片对象
        img = Image.open(image.ref).convert("RGB")
        # 将Pillow中的图片对象转换成ndarray数组
        #img = numpy.array(img)
        if i == 1 :
            img1 = img
        if i == 2 :
            img2 = img
        if i == 3:
            img3 = img

# 4.比较每个像素值
    w = True
    if img1.size != img2.size:
        w = False
    else:
        pixels1 = img1.load()
        pixels2 = img2.load()
        width, height = img1.size
        for x in range(width):
            for y in range(height):
                if pixels1[x, y] != pixels2[x, y]:
                    w = False

    print (w)
    print (img1,img2,img3)

# def hyper():
#     #from openpyxl import load_workbook
#         main_book = openpyxl.load_workbook("D:\\Users\\zhongsiyang\\test.xlsx")
#         main_sheet = main_book.active
#         print(main_sheet.cell(1, 1).value)
#         if main_sheet.cell(1,1).hyperlink:
#             print(main_sheet.cell(1, 1).hyperlink.target)
#         else:print("nothing")


# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    app = QtWidgets.QApplication(sys.argv)  # 初始化界面
    # MainWindow = QtWidgets.QMainWindow()  # 生成一个主窗口
    #
    # #MainWindow.show()  # 显示主窗口
    # page = Testy()
    # page.setupUi(MainWindow)
    # MainWindow.show()
    window = UiDiffPage()
    window.show()
    sys.exit(app.exec_())  # 在线程中退出



# See PyCharm help at https://www.jetbrains.com/help/pycharm/
